from Architecture import MainSearcherClass, MainInputClass
import os
import datetime
import time

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from Exceptions import StandartIsDone, WaitingTime, StrValidationError


def string_validator(string: str):
    if isinstance(string, str):
        pass
    else:
        raise StrValidationError("Данные в ячейке Excel были сохранены в неверном формате!")


def safe_saving(book_object, book_name, waiting_time: int):
    for _ in range(waiting_time):
        try:
            book_object.save(book_name)
            return 1
        except PermissionError:
            time.sleep(1)
    return


def connection_to_list_db():
    """Функция отвечает за своевременное обновление списка сотрудников"""

    employee_names_list = list()

    employee_book_name = "Сотрудники.xlsx"
    employee_page_name = "Список сотрудников"

    employee_book = load_workbook(employee_book_name)
    employee_page = employee_book[employee_page_name]

    for rows in range(1, employee_page.max_row + 1):
        employee_name = employee_page['A' + str(rows)].value
        if employee_name:
            employee_names_list.append(employee_name)

    return tuple(sorted(employee_names_list, key=lambda x: x.split(" ")[0]))


class SearcherClass(MainSearcherClass):

    def __init__(self, request: str):
        self.request = request

    def connection_to_db(self):
        """Функция находит и возвращает объект необходимой книги"""

        """Вначале получаем все имена файлов и папок, содержащихся в текущей
        директории"""
        path = os.getcwd()
        files = os.listdir(path)

        """Затем итерируемся по всем именам в текущей директории, извлекаем из
        них символы с 5 по 7 включительно и сравниваем их с цифрами, поступившими
        из запроса"""
        for name_of_work_book in files:
            if self.request[0:3] == name_of_work_book[4:7]:
                book = load_workbook(name_of_work_book)
                return book, name_of_work_book

    def find_page_and_carriage_position(self, work_book):
        """Функция находит и возвращает объект необходимой страницы"""

        """Вначале получаем все имена страниц в книге
        и отбираем из них только первое"""
        page_name = work_book.sheetnames[0]

        """Получив нужное нам имя листа, """
        page = work_book[page_name]

        """Теперь добудем последнюю использовавшуюся строку на листе"""
        carriage = page.max_row
        return page, carriage

    def searching_for_match(self, page_object, carriage: int):
        """Функция находит строку согласно запросу от пользователя. Помимо прочего,
        функция создает объект кэша, которых хранит в себе предыдущие проходы по
        строкам таблицы для облегчения поиска CAS-номера в дальнейшем"""

        """Создадим словарь, в котором будем хранить кэш"""
        cas_cache = {}
        catalogue_cache = {}
        name_cache = {}

        """Пройдемся по строкам таблицы, начиная с последней использовавшейся строки
        и заканчивая нулевой"""
        for row in range(carriage, 0, -1):

            """Из нужного столбца текущей строки получаем значение CAS-номера
            и, если его ещё нет в кэше, заносим его туда как ключ. В качестве
            значения используем номер строки. Это нужно для того, чтобы
            впоследствие можно было быстрее находить общий остаток для строк, 
            у которых есть CAS-номера"""
            cas_value = page_object["D" + str(row)].value
            if cas_value not in cas_cache:
                cas_cache[cas_value] = row

            """Из нужного столбца текущей строки получаем значение каталожного
            номера образца и, если его ещё нет в кэше, заносим его туда как ключ.
            В качестве значения используем номер строки. Это нужно для того,
            чтобы впоследствие можно было быстрее находить общий остаток для строк,
            у которых НЕТ CAS-номера, но есть каталожный номер"""
            catalogue_number = page_object["C" + str(row)].value
            if catalogue_number not in catalogue_cache:
                catalogue_cache[catalogue_number] = row

            """Из столбца с именем получаем значение наименования стандарта
            и заносим его туда как ключ. В качестве значения используем номер строки. 
            Это нужно для того, чтобы можно было быстрее находить общий остаток
            для строк, у которых нет ни CAS-номера, ни каталожника"""
            name_number = page_object["E" + str(row)].value
            if name_number not in name_cache:
                name_cache[name_number] = row

            """Проверяем, если цифры кода заканчиваются так же,
            как и код запроса, то мы нашли нужную строку. Возвращаем эту строку
            и кэш для дальнейшего использования"""
            if str(page_object["H" + str(row)].value) == str(self.request):
                return row, cas_cache, catalogue_cache, name_cache


class InputClass(MainInputClass):
    def __init__(self, page_object, subtractor, position_of_row: int, previous_carriage_position: int,
                 cas_cache: dict, catalogue_cache: dict, name_cache: dict, employee: str):
        self.page_object = page_object
        self.position_of_row = position_of_row
        self.previous_carriage_position = previous_carriage_position
        self.cas_cache = cas_cache
        self.catalogue_cache = catalogue_cache
        self.name_cache = name_cache
        self.employee = employee

        """Получаем номер новой строки, куда будут вноситься скопированные из нужной
        строки данные"""
        self.new_row = str(self.previous_carriage_position + 1)

        """Приводим номер строки с совпадением к строковому виду"""
        self.str_position_of_row = str(self.position_of_row)

        self.subtractor = subtractor

    def correct_input(self):
        """Функция использует:
        1) объект листа
        2) код-запрос пользователя
        3) введенную пользователем навеску
        4) номер найденной строки с совпадением,
        5) каретку позиции последней строки с данными
        6) кэш, содержащий осуществленные ранее проходы по строкам таблицы,
        чтобы внести в новую строку нужные значения полей"""

        """Проверим, не содержится ли в N-поле значения о том, что текущий выбранный
        стандарт уже израсходован. Если это так, или если значение текущего остатка во флаконе
        равно нулю, тогда возвращаем ошибку"""
        if self.page_object['N' + str(self.position_of_row)].value or \
                int(self.page_object['K' + str(self.position_of_row)].value) == 0:
            raise StandartIsDone("Стандарт уже израсходован!")

        """Создаем пары ячеек для обмена данными между ними"""
        pares_of_sheets = zip(self.page_object[('C' + self.new_row):('I' + self.new_row)][0],
                              self.page_object[('C' + self.str_position_of_row):('I' + self.str_position_of_row)][0])

        """Помещаем в новую строку извлеченные из предыдущей строки значения"""
        for new_input_row, previous_input_row in pares_of_sheets:
            new_input_row.value = previous_input_row.value

        """Проверим с помощью функции-валидатора правильность заполнения данных
        в ячейке идентификатора"""
        string_validator(self.page_object["H" + self.new_row].value)

        """Помещаем данные в L-ячейки"""
        self.page_object['L' + self.new_row].value = self.page_object['L' + self.str_position_of_row].value

        """Помещаем текущее время и дату в А-ячейки"""
        now = datetime.datetime.now()
        self.page_object['A' + self.new_row].value = now.strftime("%d-%m-%Y %H:%M")

        """Помещаем имя пользователя в ячейку"""
        self.page_object['B' + self.new_row].value = self.employee

        """Получим значение текущего остатка субстанции во флаконе"""
        previous_value_of_sample = int(self.page_object['K' + self.str_position_of_row].value)

        """Проверим - если веденная пользователем навеска меньше остатка во флаконе, тогда вернем
        значение остатка и вызовем ошибку"""
        if previous_value_of_sample < self.subtractor:
            return previous_value_of_sample

        """Проверим - не израсходовала ли взятая пользователем навеска весь остаток по флакону.
         Если израсходовала, то следует внести в N-столбец уведомление об этом"""
        if previous_value_of_sample == self.subtractor:
            self.page_object['N' + self.new_row].value = "ИЗРАСХОДОВАН"
            self.page_object['N' + self.new_row].alignment = Alignment(horizontal='center', vertical='center')

        """Сохраняем в ячейку 'остатка во флаконе' актуальный остаток"""
        self.page_object['K' + self.new_row].value = previous_value_of_sample - self.subtractor

        """Сохраняем в ячейку 'взятая навеска' взятую пользователем навеску"""
        self.page_object['J' + str(self.new_row)].value = self.subtractor

    def correct_cas_input(self):
        """Функция использует прежние параметры (аналогичные параметрам, которые
        передавались в метод correct_input), чтобы обновить данные общего остатка
        вещества для данной строки"""

        """Получаем текущий CAS-номер найденного образца"""
        cas_number = self.page_object['D' + str(self.position_of_row)].value

        """Получаем текущий каталожный номер найденного образца"""
        catalogue_number = self.page_object['C' + str(self.position_of_row)].value

        """Получаем имя найденного образца"""
        name_of_sample = self.page_object['E' + str(self.position_of_row)].value

        """Вначале проверим, существует ли в принципе CAS-номер.
        Если да, то начнем поиск по нему"""
        if cas_number != "N/A" and cas_number is not None:

            """Обращаемся к кэшу CAS-номеров - есть ли в нём значения нужного нам CAS-номера?"""
            if cas_number in self.cas_cache:
                """Если есть, тогда извлекаем из M-ячейки нужные предыдущие значения и
                в новую строку вставляем разницу между предыдущем значением и взятой
                пользователем навеской"""
                row_of_latest_cas = self.cas_cache[cas_number]
                previous_sum_value = self.page_object['M' + str(row_of_latest_cas)].value
                self.page_object['M' + str(self.new_row)].value = int(previous_sum_value) - self.subtractor
            else:
                """Если в кэше CAS-номеров не нашлось нужного значения, тогда берем значение
                общего остатка из строки с совпадением и вычитаем из него введенную пользователем 
                навеску"""
                previous_sum_value = self.page_object['M' + self.str_position_of_row].value
                self.page_object['M' + str(self.new_row)].value = int(previous_sum_value) - self.subtractor

        elif catalogue_number != "N/A" and catalogue_number is not None:
            """Если условие наличия CAS-номера не соблюдено, тогда следующим шагом постараемся
            осуществить поиск по каталожному номеру, если он присутствует"""

            """Обращаемся к кэшу каталожных номеров - есть ли в нём значения нужного нам каталожника?"""
            if catalogue_number in self.catalogue_cache:
                """Если есть, тогда извлекаем из M-ячейки нужные предыдущие значения и
                в новую строку вставляем разницу между предыдущем значением и взятой
                пользователем навеской"""
                row_of_latest_catalogue = self.catalogue_cache[catalogue_number]
                previous_sum_value = self.page_object['M' + str(row_of_latest_catalogue)].value
                self.page_object['M' + str(self.new_row)].value = int(previous_sum_value) - self.subtractor
            else:
                """Если в кэше каталожных номеров не нашлось нужного значения, тогда берем значение
                общего остатка из строки с совпадением и вычитаем из него введенную пользователем
                навеску"""
                previous_sum_value = self.page_object['M' + self.str_position_of_row].value
                self.page_object['M' + str(self.new_row)].value = int(previous_sum_value) - self.subtractor

        else:
            """Если все вышеперечисленные условия не соблюдаются, тогда осуществляем наивный поиск
            по именам"""

            """Обратимся к кэшу, содержащему имя стандарта"""
            if name_of_sample in self.name_cache:
                """Если есть, тогда извлекаем из M-ячейки нужные предыдущие значения и
                в новую строку вставляем разницу между предыдущем значением и взятой
                пользователем навеской"""
                row_of_latest_name = self.name_cache[name_of_sample]
                previous_sum_value = self.page_object['M' + str(row_of_latest_name)].value
                self.page_object['M' + str(self.new_row)].value = int(previous_sum_value) - self.subtractor
            else:
                """Если в кэше имён не нашлось нужного значения, тогда берем значение
                общего остатка из строки с совпадением и вычитаем из него введенную пользователем
                навеску"""
                previous_sum_value = self.page_object['M' + self.str_position_of_row].value
                self.page_object['M' + str(self.new_row)].value = int(previous_sum_value) - self.subtractor

        """Очистим использованные ранее кэши"""
        self.cas_cache.clear()
        self.catalogue_cache.clear()
        self.name_cache.clear()

        """Приведём все строки к единообразному виду"""
        self.page_object.row_dimensions[int(self.new_row)].height = 40
        for column in self.page_object[('A' + self.new_row):('M' + self.new_row)][0]:
            column.alignment = Alignment(horizontal='center', vertical='center')


class CorrectMistakes:
    def __init__(self, table_name, current_row: int, new_subtractor):
        self.table_name = table_name
        self.current_row = current_row

        self.table_object = load_workbook(self.table_name)
        page_name = self.table_object.sheetnames[0]
        self.page_object = self.table_object[page_name]

        self.new_subtractor = new_subtractor

    def correct_subtractor(self):
        """Функция отвечает за корректное исправление предыдущей введённой навески"""

        """Проверяем обязательное условие внесение целочисленной положительной навески.
        Иначе вызываем ошибку."""
        if self.new_subtractor <= 0:
            raise ValueError

        """Вначале очищаем поле N-столбца, убирая возможно находящиеся там
        значения о том, что стандарт уже израсходован"""
        self.page_object['N' + str(self.current_row)].value = ""

        """Затем суммируем предыдущие значения взятой навески, остатка во
        флаконе и общего остатка, чтобы получить исходные данные, из которых
        мы будем вычитать уже новую навеску"""
        previous_subtractor = int(self.page_object['J' + str(self.current_row)].value)
        previous_individual_value = int(self.page_object['K' + str(self.current_row)].value)
        previous_total_value = int(self.page_object['M' + str(self.current_row)].value)
        previous_individual_summ = int(previous_subtractor + previous_individual_value)
        previous_total_summ = int(previous_subtractor + previous_total_value)

        """Проверим также, не внёс ли пользователь навеску, которая превышает
        предыдущий остаток. Если внёс, тогда возвращаем функции-управленцу
        значение текущего максимального остатка"""
        if previous_individual_summ < self.new_subtractor:
            return previous_individual_summ

        """Если нет никаких ошибок, то тогда вносим обновленные данные в нужные поля"""
        self.page_object['J' + str(self.current_row)].value = self.new_subtractor
        self.page_object['K' + str(self.current_row)].value = previous_individual_summ - self.new_subtractor
        self.page_object['M' + str(self.current_row)].value = previous_total_summ - self.new_subtractor

        """Проверяем условие того, что стандарт не израсходован до конца.
        Если он всё же израсходован, то тогда следует поставить отметку об этом
        в N-столбец"""
        if previous_individual_summ - self.new_subtractor == 0:
            self.page_object['N' + str(self.current_row)].value = "ИЗРАСХОДОВАН"
            self.page_object['N' + str(self.current_row)].alignment = Alignment(horizontal='center', vertical='center')

        """Сохраняем изменения"""
        saving_status = safe_saving(self.table_object, self.table_name, 2)
        self.table_object.close()
        """Если не получилось сохранить, тогда нужно предупредить об этом
        пользователя, выбросив ошибку"""
        if not saving_status:
            raise WaitingTime("Книга уже кем-то используется! Сохранение невозможно")

    def delete_current_row(self):
        """Функция отвечает за полное удаление последней ошибочно внесённой строки"""

        """Удаляем последнюю строку, которая содержится в глобальном кэше"""
        self.page_object.delete_rows(self.current_row, amount=1)

        """Сохраняем изменения"""
        saving_status = safe_saving(self.table_object, self.table_name, 2)
        self.table_object.close()
        """Если не получилось сохранить, тогда нужно предупредить об этом
        пользователя, выбросив ошибку"""
        if not saving_status:
            raise WaitingTime("Книга уже кем-то используется! Сохранение невозможно")


def write_off_standart(request, employee):
    """Функция отвечает за списание пользователем флакона стандарта"""
    search = SearcherClass(request)
    book, name_of_table = search.connection_to_db()
    page, carriage = search.find_page_and_carriage_position(book)
    row, cas_cache, catalogue_cache, name_cache = search.searching_for_match(page, carriage)
    subtractor = int(page['K' + str(row)].value)

    input_class = InputClass(page, subtractor, row, carriage, cas_cache,
                             catalogue_cache, name_cache, employee)
    input_class.correct_input()
    input_class.correct_cas_input()

    """Сохраняем изменения"""
    saving_status = safe_saving(book, name_of_table, 2)
    book.close()
    """Если не получилось сохранить, тогда нужно предупредить об этом
    пользователя, выбросив ошибку"""
    if not saving_status:
        raise WaitingTime("Книга уже кем-то используется! Сохранение невозможно")
