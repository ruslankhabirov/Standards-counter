from Architecture import MainReporterClass

import datetime
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

from tkinter import filedialog


def datetime_getter(date_string):
    """Функция приводит к нормальному виду введённые пользователем
    временные интервалы"""
    date_string = date_string.replace(" ", "")
    date_list = date_string.split("-")

    for index in range(len(date_list)):
        dates = date_list[index].split(".")

        if len(dates[2]) < 4:
            dates[2] = "20" + dates[2]
        if len(dates[1]) < 2:
            dates[1] = "0" + dates[1]
        if len(dates[0]) < 2:
            dates[0] = "0" + dates[0]

        dates = ".".join(dates)
        date_list[index] = datetime.datetime.strptime(dates, '%d.%m.%Y')

    """Если пользователь ввёл более двух дат в поддерживаемом формате,
    то возбудить ошибку"""
    if len(date_list) > 2:
        raise ValueError
    if len(date_list) == 2 and date_list[0] > date_list[1]:
        date_list[0], date_list[1] = date_list[1], date_list[0]

    return date_list


def projects_list_getter(projects_string):
    return projects_string.split("; ")


def unique_not_done_data(row, data, project_name):
    """Функция, вызываемая в случае, если не было найдено совпадения внутри
    списка датафреймов (если ни один из датафреймов из списка уникальных
    идентификаторов не содержит в свою очередь совпадений по наименованию,
    уникальному идентификатору и производителю для незаконченных флаконов)"""
    data.loc[0] = None
    data.iloc[0, data.columns.get_loc(1)] = ' '.join(project_name.split(' ')[1:])
    data.iloc[0, data.columns.get_loc(2)] = row['Номер по каталогу']
    data.iloc[0, data.columns.get_loc(3)] = row['#CAS']
    data.iloc[0, data.columns.get_loc(4)] = row['Наименование стандартного образца']
    data.iloc[0, data.columns.get_loc(5)] = row['Производитель']
    data.iloc[0, data.columns.get_loc(6)] = row['Номер серии']

    if row['Номинал, мг'] == row['Остаток во флаконе, мг']:
        data.iloc[0, data.columns.get_loc(7)] = str(row['ID']) + " / " + str(row['Номинал, мг'])
        data.iloc[0, data.columns.get_loc(8)] = row['Номер серии']
        data.iloc[0, data.columns.get_loc(9)] = 1
    else:
        data.iloc[0, data.columns.get_loc(10)] = str(row['ID']) + " / " + str(row['Номинал, мг'])
        data.iloc[0, data.columns.get_loc(11)] = row['Номер серии']
        data.iloc[0, data.columns.get_loc(12)] = row['Остаток во флаконе, мг']


def unique_done_data(row, data, project_name):
    data.loc[0] = None
    data.iloc[0, data.columns.get_loc(1)] = ' '.join(project_name.split(' ')[1:])
    data.iloc[0, data.columns.get_loc(2)] = row['Номер по каталогу']
    data.iloc[0, data.columns.get_loc(3)] = row['#CAS']
    data.iloc[0, data.columns.get_loc(4)] = row['Наименование стандартного образца']
    data.iloc[0, data.columns.get_loc(5)] = row['Производитель']
    data.iloc[0, data.columns.get_loc(6)] = row['Номер серии']

    data.iloc[0, data.columns.get_loc(13)] = str(row['ID']) + " / " + str(row['Номинал, мг'])
    data.iloc[0, data.columns.get_loc(14)] = row['Номер серии']
    data.iloc[0, data.columns.get_loc(15)] = 1


def unique_not_empty_data(current_dataframe, data, expenditure, project_name):
    cas_number = str(current_dataframe.head(1)['#CAS'].values[0])
    catalogue_number = str(current_dataframe.head(1)['Номер по каталогу'].values[0])
    standard_name = str(current_dataframe.head(1)['Наименование стандартного образца'].values[0])
    manufacturer = str(current_dataframe.head(1)['Производитель'].values[0])
    serial_number = str(current_dataframe.head(1)['Номер серии'].values[0])
    standard_identificator = str(current_dataframe.head(1)['ID'].values[0])
    nominal_value = str(current_dataframe.head(1)['Номинал, мг'].values[0])

    data.loc[0] = None
    data.iloc[0, data.columns.get_loc(1)] = ' '.join(project_name.split(' ')[1:])
    data.iloc[0, data.columns.get_loc(2)] = catalogue_number
    data.iloc[0, data.columns.get_loc(3)] = cas_number
    data.iloc[0, data.columns.get_loc(4)] = standard_name
    data.iloc[0, data.columns.get_loc(5)] = manufacturer
    data.iloc[0, data.columns.get_loc(6)] = serial_number

    data.iloc[0, data.columns.get_loc(16)] = standard_identificator + " / " + nominal_value
    data.iloc[0, data.columns.get_loc(17)] = serial_number
    data.iloc[0, data.columns.get_loc(18)] = expenditure


def identical_not_done_data(row, data):
    """Функция, вызываемая в случае нахождения совпадения внутри списка
    датафреймов (если внутри списка единичных датафреймов существует такой,
    внутри котороо в свою очередь есть совпадение по наименованию, уникальному
    идентификатору и производителю для незаконченных флаконов)"""

    """Если совпадение нашлось, тогда:
     1) Проверяем, есть ли серийный номер нового образца в списке серийных номеров,
     содержащихся в поле под номером 6. Если его там нет, его следует туда занести
     2) Проверяем, является ли новый флакон вскрытым или невскрытым, и в зависимости
     от этого вносим его в правильную графу, перед этим получив номер последней
     ненулевой строки.
     3) Добавляем новую строку в датафрейм и помещаем туда, в нужные столбцы ID, номинал,
     серию и остатки по данной позиции. Перед этим проверим, есть ли ещё строка для вставки
     нового значения. Если её нет - то оную строку нужно создать, поместив туда нулевые
     значения"""
    if row['Номер серии'] not in str(data.head(1)[6].values[0]).split(', '):
        previous_data = str(data.head(1)[6].values[0]) + ', '
        new_data = str(row['Номер серии'])
        data.iloc[0, data.columns.get_loc(6)] = previous_data + new_data

    if row['Номинал, мг'] == row['Остаток во флаконе, мг']:
        checker = data.loc[pd.notnull(data[7])][7].tolist()
        if not checker:
            last_row = -1
        else:
            last_row = data.loc[pd.notnull(data[7])].tail(1).index[0]
        try:
            data.iloc[last_row + 1, data.columns.get_loc(7)] = str(row['ID']) + " / " + str(row['Номинал, мг'])
            data.iloc[last_row + 1, data.columns.get_loc(8)] = row['Номер серии']
            data.iloc[last_row + 1, data.columns.get_loc(9)] = 1
        except IndexError:
            data.loc[last_row + 1] = None
        data.iloc[last_row + 1, data.columns.get_loc(7)] = (str(row['ID']) + " / " + str(row['Номинал, мг']))
        data.iloc[last_row + 1, data.columns.get_loc(8)] = row['Номер серии']
        data.iloc[last_row + 1, data.columns.get_loc(9)] = 1
    else:
        checker = data.loc[pd.notnull(data[10])][10].tolist()
        if not checker:
            last_row = -1
        else:
            last_row = data.loc[pd.notnull(data[10])].tail(1).index[0]
        try:
            data.iloc[last_row + 1, data.columns.get_loc(10)] = str(row['ID']) + " / " + str(row['Номинал, мг'])
            data.iloc[last_row + 1, data.columns.get_loc(11)] = row['Номер серии']
            data.iloc[last_row + 1, data.columns.get_loc(12)] = row['Остаток во флаконе, мг']
        except IndexError:
            data.loc[last_row + 1] = None
        data.iloc[last_row + 1, data.columns.get_loc(10)] = (str(row['ID']) + " / " + str(row['Номинал, мг']))
        data.iloc[last_row + 1, data.columns.get_loc(11)] = row['Номер серии']
        data.iloc[last_row + 1, data.columns.get_loc(12)] = row['Остаток во флаконе, мг']


def identical_done_data(row, data):
    if row['Номер серии'] not in str(data.head(1)[6].values[0]).split(', '):
        previous_data = str(data.head(1)[6].values[0]) + ', '
        new_data = str(row['Номер серии'])
        data.iloc[0, data.columns.get_loc(6)] = previous_data + new_data

    checker = data.loc[pd.notnull(data[13])][13].tolist()

    if not checker:
        last_row = -1
    else:
        last_row = data.loc[pd.notnull(data[13])].tail(1).index[0]
    try:
        data.iloc[last_row + 1, data.columns.get_loc(13)] = str(row['ID']) + " / " + str(row['Номинал, мг'])
        data.iloc[last_row + 1, data.columns.get_loc(14)] = row['Номер серии']
        data.iloc[last_row + 1, data.columns.get_loc(15)] = 1
    except IndexError:
        data.loc[last_row + 1] = None
        data.iloc[last_row + 1, data.columns.get_loc(13)] = str(row['ID']) + " / " + str(row['Номинал, мг'])
        data.iloc[last_row + 1, data.columns.get_loc(14)] = row['Номер серии']
        data.iloc[last_row + 1, data.columns.get_loc(15)] = 1


def identical_not_empty_data(current_dataframe, data, expenditure):
    serial_number = str(current_dataframe.head(1)['Номер серии'].values[0])
    standard_identificator = str(current_dataframe.head(1)['ID'].values[0])
    nominal_value = str(current_dataframe.head(1)['Номинал, мг'].values[0])

    if serial_number not in str(data.head(1)[6].values[0]).split(', '):
        previous_data = str(data.head(1)[6].values[0]) + ', '
        data.iloc[0, data.columns.get_loc(6)] = previous_data + serial_number

    checker = data.loc[pd.notnull(data[16])][16].tolist()

    if not checker:
        last_row = -1
    else:
        last_row = data.loc[pd.notnull(data[16])].tail(1).index[0]
    try:
        data.iloc[last_row + 1, data.columns.get_loc(16)] = standard_identificator + " / " + nominal_value
        data.iloc[last_row + 1, data.columns.get_loc(17)] = serial_number
        data.iloc[last_row + 1, data.columns.get_loc(18)] = expenditure
    except IndexError:
        data.loc[last_row + 1] = None
        data.iloc[last_row + 1, data.columns.get_loc(16)] = standard_identificator + " / " + nominal_value
        data.iloc[last_row + 1, data.columns.get_loc(17)] = serial_number
        data.iloc[last_row + 1, data.columns.get_loc(18)] = expenditure


class ReporterClass(MainReporterClass):
    def __init__(self, project_name: str):
        self.project_name = project_name
        self.file_name = project_name + ".xlsx"
        self.df = pd.read_excel(self.file_name, converters={"ID": str, "Номер серии": str})
        self.df["Дата"] = pd.to_datetime(self.df["Дата"], format="%d-%m-%Y %H:%M")

    def create_done_data_frame(self, date_list: list):
        """Функция создает датафрейм из данных, собранных за определённый временной
        интервал. Возвращает объект данного датафрейма, а также список идентификаторов
        флаконов, из которых брали навески за оный срок."""

        """Проверим, какой именно список дат нам поступил - одинарный или
        двойной. Если двойной, то данные собираются за временной интервал.
        Если одинарный - то только за один день"""
        if not date_list:
            row = self.df.shape[0]
            empty_df = self.df.drop(self.df.index[[i for i in range(row)]])
            return empty_df, []

        if len(date_list) == 2:
            done_data = self.df.loc[(date_list[0] < self.df["Дата"]) & (self.df["Дата"] < date_list[1])]
            list_of_total_data = done_data.loc[pd.notnull(done_data['Взятая навеска, мг'])][
                "ID"].drop_duplicates().tolist()
            return done_data, list_of_total_data
        elif len(date_list) == 1:
            end = date_list[0] + datetime.timedelta(days=1)
            done_data = self.df.loc[(date_list[0] < self.df["Дата"]) & (self.df["Дата"] < end)]
            list_of_total_data = done_data.loc[pd.notnull(done_data['Взятая навеска, мг'])][
                "ID"].drop_duplicates().tolist()
            return done_data, list_of_total_data

    def create_not_done_data_frame(self):
        """Функция повторяет поведение той, что создаёт датафрейм израсходованных
        стандартов. За исключением того, что идентификационные номера оставшихся
        стандартов извлекаются методом исключения - для чего создается более
        широкий список израсходованных стандартов (всех, когда-либо израсходованных).
        Оставшиеся идентификационные номера и являются искомыми неизрасходованными.
        В результате возвращается датафрейм, состоящий из последних строк, в которых
        упоминается идентификационный номер неизрасходованных стандартов, а также
        список этих идентификационных номеров"""

        total_done_data_list = (self.df.loc[self.df["Статус расходования"] == "ИЗРАСХОДОВАН"])["ID"].tolist()
        not_done_data = self.df.query("ID != @total_done_data_list").groupby("ID").apply(lambda x: x.tail(1))
        if not not_done_data.empty:
            list_of_not_done_data = not_done_data["ID"].drop_duplicates().tolist()
        else:
            row = self.df.shape[0]
            empty_dataframe = self.df.drop(self.df.index[[i for i in range(row)]])
            return empty_dataframe, []
        return not_done_data, list_of_not_done_data

    def create_single_not_done_data(self, not_done_data):
        """Управляющая функция, позволяющая корректно помещать внутрь датафрейма
        значения неизрасходованных флаконов"""

        cas_unique_names = []
        catalogue_unique_names = []
        unique_names = []

        for index, row in not_done_data.iterrows():

            """Проверяем, есть ли у строки ненулевое значение поля #CAS"""
            if pd.notnull(row['#CAS']) and not (str(row['#CAS']) == "nan"):
                """Для начала проверим, есть ли дубликат данной позиции в предыдущих
                datafram-ах, содержащихся в словаре cas-датафреймов. Проверку будем
                осуществлять через поиск строки с одинаковым наименованием и одновременно
                одинаковым производителем и CAS-номером"""

                if cas_unique_names:
                    length = len(cas_unique_names)
                    """Создаем счетчик совпадений"""
                    count = 0
                    """Создаем булево выражения для проверки, не содержится ли совпадение в
                    последней итерации"""
                    last_data = cas_unique_names[-1]
                    last_iter_check = (row['Наименование стандартного образца'] == last_data.head(1)[4].values[0] and
                                       row['Производитель'] == last_data.head(1)[5].values[0] and
                                       row['#CAS'] == last_data.head(1)[3].values[0])
                    for ind in range(length):
                        count += 1
                        data = cas_unique_names[ind]
                        if (row['Наименование стандартного образца'] == data.head(1)[4].values[0] and
                                row['Производитель'] == data.head(1)[5].values[0] and
                                row['#CAS'] == data.head(1)[3].values[0]):
                            # print("Стандарт {}: Обнаружено CAS-совпадение".format(row['ID']))
                            identical_not_done_data(row, data)
                            break
                    if count == length and not last_iter_check:
                        """Если все итерации по CAS-списку прошли, и в последнем элементе этого списка не 
                        оказалось совпадений, то необходимо создать датафрейм, куда будут вноситься данные.
                        Этот датафрейм будет промежуточным и будет содержать только данные для
                        одной позиции. В качестве наименований столбцов используются числовые индексы
                        от 1 до 20 включительно"""
                        # print("Стандарт {}: CAS-совпадений нет, создан новый датафрейм".format(row['ID']))
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_not_done_data(row, new_df, self.project_name)
                        cas_unique_names.append(new_df)
                else:
                    # print("Стандарт {}: CAS-словарь пуст, создан новый датафрейм".format(row['ID']))
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_not_done_data(row, new_df, self.project_name)
                    cas_unique_names.append(new_df)

            elif pd.notnull(row['Номер по каталогу']) and not (str(row['Номер по каталогу']) == "nan"):
                if catalogue_unique_names:
                    length = len(catalogue_unique_names)
                    count = 0
                    last_data = catalogue_unique_names[-1]
                    last_iter_check = (
                            row['Наименование стандартного образца'] == str(last_data.head(1)[4].values[0]) and
                            row['Производитель'] == str(last_data.head(1)[5].values[0]) and
                            row['Номер по каталогу'] == str(last_data.head(1)[2].values[0]))
                    for ind in range(length):
                        count += 1
                        data = catalogue_unique_names[ind]
                        if (row['Наименование стандартного образца'] == str(data.head(1)[4].values[0]) and
                                row['Производитель'] == str(data.head(1)[5].values[0]) and
                                row['Номер по каталогу'] == str(data.head(1)[2].values[0])):
                            # print("Стандарт {}: Обнаружено каталог-совпадение".format(row['ID']))
                            identical_not_done_data(row, data)
                            break
                    if count == length and not last_iter_check:
                        # print("Стандарт {}: Каталог-совпадений не обнаружено, создан новый датафрейм".format(row['ID']))
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_not_done_data(row, new_df, self.project_name)
                        catalogue_unique_names.append(new_df)
                else:
                    # print("Стандарт {}: Каталог-словарь пуст, создан новый датафрейм".format(row['ID']))
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_not_done_data(row, new_df, self.project_name)
                    catalogue_unique_names.append(new_df)

            else:
                if unique_names:
                    length = len(unique_names)
                    count = 0
                    last_data = unique_names[-1]
                    last_iter_check = (
                            row['Наименование стандартного образца'] == str(last_data.head(1)[4].values[0]) and
                            row['Производитель'] == str(last_data.head(1)[5].values[0]))
                    for ind in range(length):
                        count += 1
                        data = unique_names[ind]
                        if (row['Наименование стандартного образца'] == str(data.head(1)[4].values[0]) and
                                row['Производитель'] == str(data.head(1)[5].values[0])):
                            # print("Стандарт {}: Обнаружено совпадение имени".format(row['ID']))
                            identical_not_done_data(row, data)
                            break
                    if count == length and not last_iter_check:
                        # print("Стандарт {}: Совпадений имени не обнаружено, создан новый датафрейм".format(row['ID']))
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_not_done_data(row, new_df, self.project_name)
                        unique_names.append(new_df)
                else:
                    # print("Стандарт {}: Словарь имён пуст, создан новый датафрейм".format(row['ID']))
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_not_done_data(row, new_df, self.project_name)
                    unique_names.append(new_df)

        return cas_unique_names, catalogue_unique_names, unique_names

    def empty_standards(self, done_data, cas_unique_names, catalogue_unique_names, unique_names):
        empty_standards_df = done_data.loc[done_data['Статус расходования'] == 'ИЗРАСХОДОВАН']

        for index, row in empty_standards_df.iterrows():
            if pd.notnull(row['#CAS']) and not (str(row['#CAS']) == "nan"):
                if cas_unique_names:
                    count = 0
                    length = len(cas_unique_names)
                    last_data = cas_unique_names[-1]
                    last_iter_check = (row['Наименование стандартного образца'] == last_data.head(1)[4].values[0] and
                                       row['Производитель'] == last_data.head(1)[5].values[0] and
                                       row['#CAS'] == last_data.head(1)[3].values[0])
                    for ind in range(length):
                        count += 1
                        data = cas_unique_names[ind]
                        if (data.head(1)[4].values[0] == row['Наименование стандартного образца'] and
                                data.head(1)[3].values[0] == row['#CAS'] and
                                data.head(1)[5].values[0] == row['Производитель']):
                            identical_done_data(row, data)
                            break
                    if count == length and not last_iter_check:
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_done_data(row, new_df, self.project_name)
                        cas_unique_names.append(new_df)
                else:
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_done_data(row, new_df, self.project_name)
                    cas_unique_names.append(new_df)

            elif pd.notnull(row['Номер по каталогу']) and not (str(row['Номер по каталогу']) == "nan"):
                if catalogue_unique_names:
                    count = 0
                    length = len(catalogue_unique_names)
                    last_data = catalogue_unique_names[-1]
                    last_iter_check = (
                            row['Наименование стандартного образца'] == str(last_data.head(1)[4].values[0]) and
                            row['Производитель'] == str(last_data.head(1)[5].values[0]) and
                            row['Номер по каталогу'] == str(last_data.head(1)[2].values[0]))
                    for ind in range(length):
                        count += 1
                        data = catalogue_unique_names[ind]
                        if (data.head(1)[4].values[0] == row['Наименование стандартного образца'] and
                                data.head(1)[2].values[0] == row['Номер по каталогу'] and
                                data.head(1)[5].values[0] == row['Производитель']):
                            identical_done_data(row, data)
                            break
                    if count == length and not last_iter_check:
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_done_data(row, new_df, self.project_name)
                        catalogue_unique_names.append(new_df)
                else:
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_done_data(row, new_df, self.project_name)
                    catalogue_unique_names.append(new_df)

            else:
                if unique_names:
                    count = 0
                    length = len(unique_names)
                    last_data = unique_names[-1]
                    last_iter_check = (
                            row['Наименование стандартного образца'] == str(last_data.head(1)[4].values[0]) and
                            row['Производитель'] == str(last_data.head(1)[5].values[0]))
                    for ind in range(length):
                        data = unique_names[ind]
                        if (data.head(1)[4].values[0] == row['Наименование стандартного образца'] and
                                data.head(1)[5].values[0] == row['Производитель']):
                            identical_done_data(row, data)
                            break
                    if count == length and not last_iter_check:
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_done_data(row, new_df, self.project_name)
                        unique_names.append(new_df)
                else:
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_done_data(row, new_df, self.project_name)
                    unique_names.append(new_df)

        return cas_unique_names, catalogue_unique_names, unique_names

    def not_empty_data(self, list_of_total_data, done_data, cas_unique_names, catalogue_unique_names, unique_names):

        done_data = done_data.query("ID == @list_of_total_data")
        done_data = done_data.loc[pd.notnull(done_data['Взятая навеска, мг'])]

        for number_of_standard in list_of_total_data:
            current_dataframe = done_data.loc[done_data['ID'] == number_of_standard]
            expenditure = current_dataframe['Взятая навеска, мг'].astype(int).sum()
            standard_name = str(current_dataframe.head(1)['Наименование стандартного образца'].values[0])
            standard_catalogue_name = str(current_dataframe.head(1)['Номер по каталогу'].values[0])
            standard_cas_name = str(current_dataframe.head(1)['#CAS'].values[0])
            standard_manufacturer = str(current_dataframe.head(1)['Производитель'].values[0])

            """if isinstance(standard_cas_name, float):
                print("Поймал! Тип cas-номера {}: {}".format(standard_cas_name, type(standard_cas_name)))
            if isinstance(standard_cas_name, str):
                print("Поймал! Тип cas-номера {}: {}".format(standard_cas_name, type(standard_cas_name)))"""

            if pd.notnull(standard_cas_name) and not (standard_cas_name == "nan"):
                if cas_unique_names:
                    count = 0
                    length = len(cas_unique_names)
                    last_data = cas_unique_names[-1]
                    last_iter_check = (standard_name == last_data.head(1)[4].values[0] and
                                       standard_manufacturer == last_data.head(1)[5].values[0] and
                                       standard_cas_name == last_data.head(1)[3].values[0])
                    for ind in range(length):
                        count += 1
                        data = cas_unique_names[ind]
                        if (data.head(1)[3].values[0] == standard_cas_name and
                                data.head(1)[4].values[0] == standard_name and
                                data.head(1)[5].values[0] == standard_manufacturer):
                            # print("Стандарт {}: Обнаружено CAS-совпадение".format(number_of_standard))
                            identical_not_empty_data(current_dataframe, data, expenditure)
                            break
                    if count == length and not last_iter_check:
                        # print("Стандарт {}: CAS-совпадений не обнаружено, создан новый df".format(number_of_standard))
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_not_empty_data(current_dataframe, new_df, expenditure, self.project_name)
                        cas_unique_names.append(new_df)
                else:
                    # print("Стандарт {}: CAS-список пуст, создан новый df".format(number_of_standard))
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_not_empty_data(current_dataframe, new_df, expenditure, self.project_name)
                    cas_unique_names.append(new_df)

            elif pd.notnull(standard_catalogue_name) and not (standard_catalogue_name == "nan"):
                if catalogue_unique_names:
                    count = 0
                    length = len(catalogue_unique_names)
                    last_data = catalogue_unique_names[-1]
                    last_iter_check = (standard_name == str(last_data.head(1)[4].values[0]) and
                                       standard_manufacturer == str(last_data.head(1)[5].values[0]) and
                                       standard_catalogue_name == str(last_data.head(1)[2].values[0]))
                    for ind in range(length):
                        count += 1
                        data = catalogue_unique_names[ind]
                        if (data.head(1)[2].values[0] == standard_catalogue_name and
                                data.head(1)[4].values[0] == standard_name and
                                data.head(1)[5].values[0] == standard_manufacturer):
                            # print("Стандарт {}: Обнаружено каталог-совпадение".format(number_of_standard))
                            identical_not_empty_data(current_dataframe, data, expenditure)
                            break
                    if count == length and not last_iter_check:
                        # print("Стандарт {}: Каталог-совпадений не обнаружено, создан новый df".format(number_of_standard))
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_not_empty_data(current_dataframe, new_df, expenditure, self.project_name)
                        catalogue_unique_names.append(new_df)
                else:
                    # print("Стандарт {}: Каталог-список пуст, создан новый df".format(number_of_standard))
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_not_empty_data(current_dataframe, new_df, expenditure, self.project_name)
                    catalogue_unique_names.append(new_df)

            else:
                if unique_names:
                    count = 0
                    length = len(unique_names)
                    last_data = unique_names[-1]
                    last_iter_check = (standard_name == str(last_data.head(1)[4].values[0]) and
                                       standard_manufacturer == str(last_data.head(1)[5].values[0]))
                    for ind in range(length):
                        count += 1
                        data = unique_names[ind]
                        if (data.head(1)[4].values[0] == standard_name and
                                data.head(1)[5].values[0] == standard_manufacturer):
                            # print("Стандарт {}: Обнаружено совпадение имени".format(number_of_standard))
                            identical_not_empty_data(current_dataframe, data, expenditure)
                            break
                    if count == length and not last_iter_check:
                        # print("Стандарт {}: Совпадений имени не обнаружено, создан новый df".format(number_of_standard))
                        new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                        unique_not_empty_data(current_dataframe, new_df, expenditure, self.project_name)
                        unique_names.append(new_df)
                else:
                    # print("Стандарт {}: Список имён пуст, создан новый df".format(number_of_standard))
                    new_df = pd.DataFrame(columns=[i for i in range(1, 21)])
                    unique_not_empty_data(current_dataframe, new_df, expenditure, self.project_name)
                    unique_names.append(new_df)

        return cas_unique_names, catalogue_unique_names, unique_names

    def add_balance(self, cas_unique_names, catalogue_unique_names, unique_names):

        list_of_balance = [cas_unique_names, catalogue_unique_names, unique_names]

        for balance in list_of_balance:
            for data in balance:

                if pd.notnull(data.iloc[0, 2]) and not str(data.iloc[0, 2]) == "nan":
                    current_cas = data.iloc[0, 2]
                    last_balance = (self.df.loc[self.df['#CAS'] ==
                                                current_cas].tail(1)['Общий остаток'].values[0])
                    last_irreducible_balance = (self.df.loc[self.df['#CAS'] ==
                                                            current_cas].tail(1)['Неснижаемый остаток'].values[0])
                    data.iloc[0, 18] = last_balance
                    if last_irreducible_balance:
                        data.iloc[0, 19] = last_irreducible_balance

                elif pd.notnull(data.iloc[0, 1]) and not str(data.iloc[0, 1]) == "nan":
                    current_catalogue = data.iloc[0, 1]
                    last_balance = (self.df.loc[self.df['Номер по каталогу'] ==
                                                current_catalogue].tail(1)['Общий остаток'].values[0])
                    last_irreducible_balance = (self.df.loc[self.df['Номер по каталогу'] ==
                                                            current_catalogue].tail(1)['Неснижаемый остаток'].values[0])
                    data.iloc[0, 18] = last_balance
                    if last_irreducible_balance:
                        data.iloc[0, 19] = last_irreducible_balance

                else:
                    current_name = data.iloc[0, 3]
                    last_balance = (self.df.loc[self.df['Наименование стандартного образца'] ==
                                                current_name].tail(1)['Общий остаток'].values[0])
                    last_irreducible_balance = (self.df.loc[self.df['Наименование стандартного образца'] ==
                                                            current_name].tail(1)['Неснижаемый остаток'].values[0])
                    data.iloc[0, 18] = last_balance
                    if last_irreducible_balance:
                        data.iloc[0, 19] = last_irreducible_balance

        return cas_unique_names, catalogue_unique_names, unique_names


def add_summ(data):
    checker_of_not_empty = data.loc[pd.notnull(data[9])][9].tolist()
    checker_of_opened = data.loc[pd.notnull(data[12])][12].tolist()
    checker_of_fully_ended = data.loc[pd.notnull(data[15])][15].tolist()
    checker_of_not_fully_ended = data.loc[pd.notnull(data[18])][18].tolist()

    check_list = [checker_of_not_empty, checker_of_opened,
                  checker_of_fully_ended, checker_of_not_fully_ended]
    numbers = [9, 12, 15, 18]
    total_list = zip(numbers, check_list)

    for num, status in total_list:
        if status:
            last_row = int(data.loc[pd.notnull(data[num])].tail(1).index[0])
            try:
                data.iloc[last_row + 1, data.columns.get_loc(num - 1)] = "Итого:"
                expenditure = data.loc[pd.notnull(data[num])][num].astype(int).sum()
                data.iloc[last_row + 1, data.columns.get_loc(num)] = expenditure
            except IndexError:
                data.loc[last_row + 1] = None
                data.iloc[last_row + 1, data.columns.get_loc(num - 1)] = "Итого:"
                expenditure = data.loc[pd.notnull(data[num])][num].astype(int).sum()
                data.iloc[last_row + 1, data.columns.get_loc(num)] = expenditure


def add_total_summ(cas_unique_names, catalogue_unique_names, unique_names):
    if cas_unique_names:
        for data in cas_unique_names:
            add_summ(data)

    if catalogue_unique_names:
        for data in catalogue_unique_names:
            add_summ(data)

    if unique_names:
        for data in unique_names:
            add_summ(data)

    return cas_unique_names, catalogue_unique_names, unique_names


def merge_dataframes(cas_unique_names, catalogue_unique_names, unique_names):
    cas_unique_names = sorted(cas_unique_names, key=lambda x: x.iloc[0, 2])
    catalogue_unique_names = sorted(catalogue_unique_names, key=lambda x: x.iloc[0, 1])
    unique_names = sorted(unique_names, key=lambda x: x.iloc[0, 3])

    total_list = cas_unique_names + catalogue_unique_names + unique_names

    total_df = pd.DataFrame(columns=[i for i in range(1, 21)])

    for data in total_list:
        last_row_index = total_df.shape[0]
        for index, data_row in data.iterrows():
            total_df.loc[last_row_index] = data_row
            last_row_index += 1

    return total_df


def summary_dataframe(data_list):
    data_list = sorted(data_list, key=lambda x: x.iloc[0, 0])
    total_df = pd.DataFrame(columns=[i for i in range(1, 21)])
    for data in data_list:
        last_row_index = total_df.shape[0]
        for index, data_row in data.iterrows():
            total_df.loc[last_row_index] = data_row
            last_row_index += 1

    return total_df


def create_excel_file(data, initial_file_name):
    for _ in range(3):
        data.loc[-1] = None
        data.index = data.index + 1
        data = data.sort_index()

    file = filedialog.asksaveasfile(initialfile=initial_file_name, filetypes=[('excel file', '.xlsx')])

    if not file.name.endswith(".xlsx"):
        name = file.name + ".xlsx"
    else:
        name = file.name

    data.to_excel(name)

    book = load_workbook(name)
    page = book.active

    page.delete_cols(1)
    page.delete_rows(1)

    last_row = page.max_row
    letters_list = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L",
                    "M", "N", "O", "P", "Q", "R", "S", "T"]
    for ind in range(1, last_row + 1):
        page.row_dimensions[ind].height = 40
        for letter in letters_list:
            page[letter + str(ind)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    page.merge_cells('A1:A3')
    page['A1'].alignment = Alignment(horizontal='center', vertical='center')
    page['A1'].value = "Проект"
    page.column_dimensions["A"].width = 20
    page['A1'].font = Font(size=14)

    page.merge_cells('B1:B3')
    page['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['B1'].value = "Номер по каталогу"
    page.column_dimensions["B"].width = 20
    page['B1'].font = Font(size=14)

    page.merge_cells('C1:C3')
    page['C1'].alignment = Alignment(horizontal='center', vertical='center')
    page['C1'].value = "#CAS"
    page.column_dimensions["C"].width = 15
    page['C1'].font = Font(size=14)

    page.merge_cells('D1:D3')
    page['D1'].alignment = Alignment(horizontal='center', vertical='center')
    page['D1'].value = "Наименование стандартного образца"
    page.column_dimensions["D"].width = 75
    page['D1'].font = Font(size=14)

    page.merge_cells('E1:E3')
    page['E1'].alignment = Alignment(horizontal='center', vertical='center')
    page['E1'].value = "Производитель"
    page.column_dimensions["E"].width = 20
    page['E1'].font = Font(size=14)

    page.merge_cells('F1:F3')
    page['F1'].alignment = Alignment(horizontal='center', vertical='center')
    page['F1'].value = "Номера серий"
    page.column_dimensions["F"].width = 20
    page['F1'].font = Font(size=14)

    page.merge_cells('G1:L1')
    page['G1'].alignment = Alignment(horizontal='center', vertical='center')
    page['G1'].value = "Остатки"
    page.column_dimensions["G"].width = 15
    page.column_dimensions["H"].width = 15
    page.column_dimensions["I"].width = 15
    page.column_dimensions["J"].width = 15
    page.column_dimensions["K"].width = 15
    page.column_dimensions["L"].width = 15
    page['G1'].font = Font(size=14)

    page.merge_cells('G2:I2')
    page['G2'].alignment = Alignment(horizontal='center', vertical='center')
    page['G2'].value = "Не вскрытые флаконы"
    page['G2'].font = Font(size=14)

    page.merge_cells('J2:L2')
    page['J2'].alignment = Alignment(horizontal='center', vertical='center')
    page['J2'].value = "Вскрытые флаконы"
    page['J2'].font = Font(size=14)

    page['G3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['G3'].value = "ID / номинал"
    page['G3'].font = Font(size=14)

    page['H3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['H3'].value = "Серия"
    page['H3'].font = Font(size=14)

    page['I3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['I3'].value = "Кол-во флаконов"
    page['I3'].font = Font(size=14)

    page['J3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['J3'].value = "ID / номинал"
    page['J3'].font = Font(size=14)

    page['K3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['K3'].value = "Серия"
    page['K3'].font = Font(size=14)

    page['L3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['L3'].value = "Остаток во флаконе"
    page['L3'].font = Font(size=14)

    page.merge_cells('M1:R1')
    page['M1'].alignment = Alignment(horizontal='center', vertical='center')
    page['M1'].value = "Расходы"
    page.column_dimensions["M"].width = 15
    page.column_dimensions["N"].width = 15
    page.column_dimensions["O"].width = 15
    page.column_dimensions["P"].width = 15
    page.column_dimensions["Q"].width = 15
    page.column_dimensions["R"].width = 15
    page['M1'].font = Font(size=14)

    page.merge_cells('M2:O2')
    page['M2'].alignment = Alignment(horizontal='center', vertical='center')
    page['M2'].value = "Целые флаконы"
    page['M2'].font = Font(size=14)

    page.merge_cells('P2:R2')
    page['P2'].alignment = Alignment(horizontal='center', vertical='center')
    page['P2'].value = "Расходы в мг по флаконам"
    page['P2'].font = Font(size=14)

    page['M3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['M3'].value = "ID / номинал"
    page['M3'].font = Font(size=14)

    page['N3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['N3'].value = "Серия"
    page['N3'].font = Font(size=14)

    page['O3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['O3'].value = "Кол-во флаконов"
    page['O3'].font = Font(size=14)

    page['P3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['P3'].value = "ID / номинал"
    page['P3'].font = Font(size=14)

    page['Q3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['Q3'].value = "Серия"
    page['Q3'].font = Font(size=14)

    page['R3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['R3'].value = "Расход по флаконам"
    page['R3'].font = Font(size=14)

    page.merge_cells('S1:S3')
    page['S1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['S1'].value = "Общий остаток по веществу"
    page['S1'].font = Font(size=14)
    page.column_dimensions["S"].width = 15

    page.merge_cells('T1:T3')
    page['T1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    page['T1'].value = "Неснижаемый остаток по веществу"
    page['T1'].font = Font(size=14)
    page.column_dimensions["T"].width = 15

    for row in range(1, 4):
        for letter in letters_list:
            page[letter + str(row)].fill = PatternFill(fill_type='solid', start_color="1F497D", end_color="1F497D")
            page[letter + str(row)].font = Font(color="FFFFFF")
            page[letter + str(row)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                    right=Side(border_style='thin', color='FF000000'),
                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                    left=Side(border_style='thin', color='FF000000'))

    right_thin_border = Border(right=Side(style='thin'))
    bottom_thin_border = Border(bottom=Side(style='thin'))
    bottom_right_thin_border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    bottom_left_thin_border = Border(bottom=Side(style='thin'), left=Side(style='thin'))
    bottom_left_right_thin_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    count = 4

    def oreder_cell_border(current_page, short_letters_list, current_row):
        current_page[short_letters_list[0] + str(current_row)].border = bottom_left_thin_border
        current_page[short_letters_list[1] + str(current_row)].border = bottom_thin_border
        current_page[short_letters_list[2] + str(current_row)].border = bottom_right_thin_border

    while count <= last_row:
        for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'I', 'L', 'O', 'R', 'S', 'T']:
            page[letter + str(count)].border = right_thin_border

        if page['A' + str(count)].value:
            count += 1
        else:
            start = count
            end = count
            for _ in range(start, last_row):
                if not page['A' + str(end + 1)].value:
                    end += 1
                else:
                    break
            for cell in page[('A' + str(end)):('F' + str(end))][0]:
                cell.border = bottom_right_thin_border
            for list_of_letters in [['G', 'H', 'I'], ['J', 'K', 'L'], ['M', 'N', 'O'], ['P', 'Q', 'R']]:
                oreder_cell_border(page, list_of_letters, str(end))
            page['S' + str(end)].border = bottom_left_right_thin_border
            page['T' + str(end)].border = bottom_left_right_thin_border

            """Сгруппируем строки"""
            page.row_dimensions.group(start=start, end=end, hidden=True)
            count += 1

    """Сгруппируем столбцы"""
    page.column_dimensions.group('G', 'R', hidden=True)

    """Установим правило - подсвечивать красным ячейки, где значения общего остатка
    меньше значений неснижаемого остатка"""

    red_fill = PatternFill(start_color='AA0000', end_color='AA0000', fill_type='solid')

    for row_number in range(4, page.max_row):
        minimal_level = page['T' + str(row_number)].value
        if minimal_level:
            rule = CellIsRule(operator='lessThan', formula=['T$' + str(row_number)], stopIfTrue=True, fill=red_fill)
            page.conditional_formatting.add('S' + str(row_number), rule)

    book.save(name)
    book.close()


